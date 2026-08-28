from io import BytesIO
from datetime import datetime
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


TITLE_FILL = "1F4E78"
HEADER_FILL = "D9EAF7"
SUB_FILL = "EAF2F8"
RED_FILL = "FDE9E7"
ORANGE_FILL = "FCE4D6"
GREEN_FILL = "E2F0D9"

WHITE = "FFFFFF"
DARK = "1F1F1F"
RED = "C00000"
ORANGE = "C65911"
BLUE = "1F4E78"
GREEN = "548235"

THIN_GRAY = Side(
    style="thin",
    color="B7C0C8"
)

DEFAULT_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY
)


# Excel/OpenXML 셀에 사용할 수 없는 제어문자 제거
# (NULL 바이트 등 파일 업로드 우회 테스트용 문자열이 포함된 경우 openpyxl 저장 오류 방지)
EXCEL_ILLEGAL_CHAR_RE = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F]"
)


def sanitize_excel_text(value):
    if value is None:
        return value

    if not isinstance(value, str):
        return value

    # 사람이 보고 원문에 제어문자가 있었다는 사실을 알 수 있도록 치환문자 사용
    return EXCEL_ILLEGAL_CHAR_RE.sub("�", value)


def safe_value(value, default="-"):
    if value is None:
        return default

    text = sanitize_excel_text(
        str(value)
    ).strip()

    if not text:
        return default

    return text



def clean_ai_report_text(value):
    """
    AI 분석 Markdown을 XLSX에서 읽기 쉬운 일반 텍스트로 정리합니다.
    내용은 유지하고 화면용 Markdown 기호/구분선만 제거합니다.
    """
    text = safe_value(value, "")

    if not text:
        return ""

    text = sanitize_excel_text(text)

    # Markdown 구분선 제거
    text = re.sub(r"(?m)^\s*[-*_]{3,}\s*$", "", text)

    # 제목 / 강조 / 인라인 코드 기호 제거
    text = re.sub(r"(?m)^\s{0,3}#{1,6}\s*", "", text)
    text = text.replace("**", "")
    text = text.replace("__", "")
    text = text.replace("`", "")

    # 불필요한 공백 정리
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()


def normalize_comparison_data(comparison_data):
    if comparison_data is None:
        return []

    if hasattr(comparison_data, "to_dict"):
        try:
            return comparison_data.to_dict("records")
        except Exception:
            pass

    if isinstance(comparison_data, list):
        return comparison_data

    return []


def style_title(
    sheet,
    cell_range,
    title
):
    sheet.merge_cells(cell_range)

    first_cell = cell_range.split(":")[0]
    cell = sheet[first_cell]

    cell.value = title

    cell.fill = PatternFill(
        "solid",
        fgColor=TITLE_FILL
    )

    cell.font = Font(
        bold=True,
        color=WHITE,
        size=16
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    sheet.row_dimensions[
        cell.row
    ].height = 28


def style_header_row(
    sheet,
    row,
    start_col,
    end_col
):
    for col in range(
        start_col,
        end_col + 1
    ):
        cell = sheet.cell(
            row=row,
            column=col
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_FILL
        )

        cell.font = Font(
            bold=True,
            color=DARK
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = DEFAULT_BORDER


def style_body_range(
    sheet,
    start_row,
    end_row,
    start_col,
    end_col
):
    if end_row < start_row:
        return

    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            cell.border = DEFAULT_BORDER


def apply_status_style(cell):
    value = str(
        cell.value
    ).strip()

    if value == "취약":
        cell.fill = PatternFill(
            "solid",
            fgColor=RED_FILL
        )
        cell.font = Font(
            bold=True,
            color=RED
        )

    elif value == "양호":
        cell.fill = PatternFill(
            "solid",
            fgColor=SUB_FILL
        )
        cell.font = Font(
            bold=True,
            color=BLUE
        )


def apply_risk_style(cell):
    value = str(
        cell.value
    ).strip()

    if value == "높음":
        cell.fill = PatternFill(
            "solid",
            fgColor=RED_FILL
        )
        cell.font = Font(
            bold=True,
            color=RED
        )

    elif value == "중간":
        cell.fill = PatternFill(
            "solid",
            fgColor=ORANGE_FILL
        )
        cell.font = Font(
            bold=True,
            color=ORANGE
        )


def set_column_widths(
    sheet,
    widths
):
    for column, width in widths.items():
        sheet.column_dimensions[
            column
        ].width = width



def generate_xlsx_report(
    results,
    comparison_data=None,
    hf_result=None,
    ai_analysis=None,
    actual_file_reports=None,
    target_name="공공 민원포털 웹 서비스"
):
    """
    사람이 읽는 보고서 용도의 XLSX.

    핵심 방향
    - 요약 시트는 빠르게 훑어볼 수 있게 압축
    - 상세 시트는 한 취약점당 하나의 카드 블록으로 분리
    - 긴 문자열 때문에 한 행이 지나치게 커져 휠 한 번에 화면이 크게 튀는 현상 방지
    - 긴 내용은 넓게 병합된 셀에 배치하고 행 높이는 일정 범위로 제한
    """
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    report_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total = len(results)
    vulnerable = sum(
        1 for item in results
        if item.get("status") == "취약"
    )
    safe = sum(
        1 for item in results
        if item.get("status") == "양호"
    )
    na = sum(
        1 for item in results
        if item.get("status") == "N/A"
    )

    def fill(cell, color):
        cell.fill = PatternFill(
            "solid",
            fgColor=color
        )

    def set_all_border(sheet, cell_range):
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = DEFAULT_BORDER

    def merged_value(
        sheet,
        cell_range,
        value,
        *,
        bg=None,
        bold=False,
        color=DARK,
        size=10,
        horizontal="left",
        vertical="top"
    ):
        sheet.merge_cells(cell_range)
        first = cell_range.split(":")[0]
        cell = sheet[first]
        cell.value = safe_value(value)
        if bg:
            fill(cell, bg)
        cell.font = Font(
            bold=bold,
            color=color,
            size=size
        )
        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=True
        )
        set_all_border(sheet, cell_range)
        return cell

    def short_text(value, limit=150):
        text = safe_value(value)
        if len(text) <= limit:
            return text
        return text[:limit].rstrip() + " ..."

    def label_cell(sheet, row, label, end_row=None):
        end_row = end_row or row
        if end_row > row:
            sheet.merge_cells(
                start_row=row,
                start_column=1,
                end_row=end_row,
                end_column=1
            )
        cell = sheet.cell(row=row, column=1, value=label)
        fill(cell, SUB_FILL)
        cell.font = Font(
            bold=True,
            color=BLUE
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        for r in range(row, end_row + 1):
            sheet.cell(r, 1).border = DEFAULT_BORDER
        return cell

    def status_risk_meta(sheet, row, item):
        labels = [
            ("판정", safe_value(item.get("status"))),
            ("위험도", safe_value(item.get("risk"))),
            ("진단 확실성", safe_value(item.get("confidence")))
        ]

        start_cols = [1, 4, 7]

        for (label, value), start_col in zip(labels, start_cols):
            label_cell_obj = sheet.cell(
                row=row,
                column=start_col,
                value=label
            )
            fill(label_cell_obj, SUB_FILL)
            label_cell_obj.font = Font(
                bold=True,
                color=BLUE
            )
            label_cell_obj.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            label_cell_obj.border = DEFAULT_BORDER

            sheet.merge_cells(
                start_row=row,
                start_column=start_col + 1,
                end_row=row,
                end_column=start_col + 2
            )

            value_cell = sheet.cell(
                row=row,
                column=start_col + 1,
                value=value
            )
            value_cell.font = Font(
                bold=True,
                color=DARK
            )
            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            for c in range(start_col + 1, start_col + 3):
                sheet.cell(row, c).border = DEFAULT_BORDER

        apply_status_style(
            sheet.cell(row, 2)
        )
        apply_risk_style(
            sheet.cell(row, 5)
        )

    # ===============================================================
    # 1) 진단 요약
    # ===============================================================
    summary_sheet = workbook.create_sheet(
        "진단 요약"
    )

    style_title(
        summary_sheet,
        "A1:H1",
        "AI 기반 공공 민원포털 취약점 자동 진단 결과"
    )

    # 기본 정보 카드
    summary_sheet["A3"] = "진단 대상"
    summary_sheet["A4"] = "보고서 생성 시각"
    summary_sheet["A5"] = "진단 방식"

    for row in range(3, 6):
        cell = summary_sheet.cell(row, 1)
        fill(cell, SUB_FILL)
        cell.font = Font(
            bold=True,
            color=BLUE
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = DEFAULT_BORDER

    for row, value in [
        (3, target_name),
        (4, report_time),
        (5, "Python 기반 자동 진단 + AI 기반 결과 분석")
    ]:
        summary_sheet.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=4
        )
        cell = summary_sheet.cell(
            row,
            2,
            safe_value(value)
        )
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )
        for col in range(2, 5):
            summary_sheet.cell(row, col).border = DEFAULT_BORDER

    # KPI
    kpi_headers = [
        "전체 진단",
        "취약",
        "양호",
        "N/A"
    ]
    kpi_values = [
        total,
        vulnerable,
        safe,
        na
    ]

    for col, header in enumerate(
        kpi_headers,
        start=5
    ):
        cell = summary_sheet.cell(
            3,
            col,
            header
        )
        fill(cell, HEADER_FILL)
        cell.font = Font(
            bold=True,
            color=DARK
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = DEFAULT_BORDER

    for col, value in enumerate(
        kpi_values,
        start=5
    ):
        cell = summary_sheet.cell(
            4,
            col,
            value
        )
        cell.font = Font(
            bold=True,
            size=15,
            color=BLUE
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = DEFAULT_BORDER

    summary_sheet["E5"] = (
        "취약률"
    )
    fill(summary_sheet["E5"], HEADER_FILL)
    summary_sheet["E5"].font = Font(
        bold=True
    )
    summary_sheet["E5"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    summary_sheet["E5"].border = DEFAULT_BORDER

    summary_sheet.merge_cells(
        "F5:H5"
    )
    rate_cell = summary_sheet["F5"]
    rate_cell.value = (
        vulnerable / total
        if total else 0
    )
    rate_cell.number_format = "0.0%"
    rate_cell.font = Font(
        bold=True,
        size=13,
        color=RED if vulnerable else GREEN
    )
    rate_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    for col in range(6, 9):
        summary_sheet.cell(5, col).border = DEFAULT_BORDER

    # 빠르게 훑는 요약 목록
    headers = [
        "No.",
        "취약점",
        "판정",
        "위험도",
        "탐지 내용 요약"
    ]

    for index, value in enumerate(
        headers,
        start=1
    ):
        summary_sheet.cell(
            row=8,
            column=index,
            value=value
        )

    style_header_row(
        summary_sheet,
        8,
        1,
        5
    )

    current_row = 9

    for idx, item in enumerate(
        results,
        start=1
    ):
        values = [
            idx,
            safe_value(
                item.get("vulnerability")
            ),
            safe_value(
                item.get("status")
            ),
            safe_value(
                item.get("risk")
            ),
            short_text(
                item.get("evidence"),
                165
            )
        ]

        for column, value in enumerate(
            values,
            start=1
        ):
            cell = summary_sheet.cell(
                row=current_row,
                column=column,
                value=value
            )
            cell.border = DEFAULT_BORDER
            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if column in [1, 3, 4]
                    else "left"
                ),
                vertical="center",
                wrap_text=True
            )

        apply_status_style(
            summary_sheet.cell(
                current_row,
                3
            )
        )
        apply_risk_style(
            summary_sheet.cell(
                current_row,
                4
            )
        )

        summary_sheet.row_dimensions[
            current_row
        ].height = 34

        current_row += 1

    summary_sheet.freeze_panes = "A9"
    summary_sheet.auto_filter.ref = (
        f"A8:E{current_row - 1}"
    )

    set_column_widths(
        summary_sheet,
        {
            "A": 7,
            "B": 39,
            "C": 12,
            "D": 12,
            "E": 72,
            "F": 14,
            "G": 14,
            "H": 14
        }
    )

    # ===============================================================
    # 2) 상세 진단 결과 - 카드형 블록
    # ===============================================================
    detail_sheet = workbook.create_sheet(
        "상세 진단 결과"
    )

    style_title(
        detail_sheet,
        "A1:I1",
        "상세 진단 결과"
    )

    detail_sheet.merge_cells(
        "A3:I3"
    )
    guide_cell = detail_sheet["A3"]
    guide_cell.value = (
        "각 취약점은 하나의 블록으로 분리되어 있습니다. "
        "긴 입력값·근거·대응방안은 넓은 병합 셀에 배치하여 "
        "한 행이 과도하게 커지지 않도록 구성했습니다."
    )
    guide_cell.fill = PatternFill(
        "solid",
        fgColor="F3F6F9"
    )
    guide_cell.font = Font(
        italic=True,
        color="5B6570"
    )
    guide_cell.alignment = Alignment(
        vertical="center",
        wrap_text=True
    )
    detail_sheet.row_dimensions[
        3
    ].height = 28

    current_row = 5

    for index, item in enumerate(
        results,
        start=1
    ):
        # 제목 행
        detail_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=9
        )

        title_cell = detail_sheet.cell(
            current_row,
            1,
            f"{index:02d}. "
            f"{safe_value(item.get('vulnerability'))}"
        )
        fill(title_cell, TITLE_FILL)
        title_cell.font = Font(
            bold=True,
            color=WHITE,
            size=11
        )
        title_cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

        for col in range(1, 10):
            detail_sheet.cell(
                current_row,
                col
            ).border = DEFAULT_BORDER

        detail_sheet.row_dimensions[
            current_row
        ].height = 27

        # 판정 / 위험도 / 확실성
        status_risk_meta(
            detail_sheet,
            current_row + 1,
            item
        )
        detail_sheet.row_dimensions[
            current_row + 1
        ].height = 24

        # 진단 대상
        label_cell(
            detail_sheet,
            current_row + 2,
            "진단 대상"
        )
        detail_sheet.merge_cells(
            start_row=current_row + 2,
            start_column=2,
            end_row=current_row + 2,
            end_column=9
        )
        target_cell = detail_sheet.cell(
            current_row + 2,
            2,
            safe_value(
                item.get("parameter")
            )
        )
        target_cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )
        for col in range(2, 10):
            detail_sheet.cell(
                current_row + 2,
                col
            ).border = DEFAULT_BORDER
        detail_sheet.row_dimensions[
            current_row + 2
        ].height = 30

        # 테스트 입력값
        label_cell(
            detail_sheet,
            current_row + 3,
            "테스트 입력값"
        )
        detail_sheet.merge_cells(
            start_row=current_row + 3,
            start_column=2,
            end_row=current_row + 3,
            end_column=9
        )
        payload_cell = detail_sheet.cell(
            current_row + 3,
            2,
            safe_value(
                item.get("payload")
            )
        )
        payload_cell.font = Font(
            name="Consolas",
            size=9,
            color="3B3B3B"
        )
        payload_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )
        for col in range(2, 10):
            detail_sheet.cell(
                current_row + 3,
                col
            ).border = DEFAULT_BORDER
        detail_sheet.row_dimensions[
            current_row + 3
        ].height = 44

        # 탐지 내용
        label_cell(
            detail_sheet,
            current_row + 4,
            "탐지 내용"
        )
        detail_sheet.merge_cells(
            start_row=current_row + 4,
            start_column=2,
            end_row=current_row + 4,
            end_column=9
        )
        evidence_cell = detail_sheet.cell(
            current_row + 4,
            2,
            safe_value(
                item.get("evidence")
            )
        )
        evidence_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )
        for col in range(2, 10):
            detail_sheet.cell(
                current_row + 4,
                col
            ).border = DEFAULT_BORDER
        detail_sheet.row_dimensions[
            current_row + 4
        ].height = 46

        # 판단 근거
        label_cell(
            detail_sheet,
            current_row + 5,
            "판단 근거"
        )
        detail_sheet.merge_cells(
            start_row=current_row + 5,
            start_column=2,
            end_row=current_row + 5,
            end_column=9
        )
        reason_cell = detail_sheet.cell(
            current_row + 5,
            2,
            safe_value(
                item.get("reason")
            )
        )
        reason_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )
        for col in range(2, 10):
            detail_sheet.cell(
                current_row + 5,
                col
            ).border = DEFAULT_BORDER
        detail_sheet.row_dimensions[
            current_row + 5
        ].height = 46

        # 대응방안
        label_cell(
            detail_sheet,
            current_row + 6,
            "대응방안"
        )
        detail_sheet.merge_cells(
            start_row=current_row + 6,
            start_column=2,
            end_row=current_row + 6,
            end_column=9
        )
        rec_cell = detail_sheet.cell(
            current_row + 6,
            2,
            safe_value(
                item.get("recommendation")
            )
        )
        rec_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )
        for col in range(2, 10):
            detail_sheet.cell(
                current_row + 6,
                col
            ).border = DEFAULT_BORDER
        detail_sheet.row_dimensions[
            current_row + 6
        ].height = 46

        # 블록 간 간격
        detail_sheet.row_dimensions[
            current_row + 7
        ].height = 9

        current_row += 8


    # ---------------------------------------------------------------
    # 실제 업로드 파일 자체 분석 결과
    # 대시보드의 "실제 업로드 파일 진단" 결과를 같은 카드 디자인으로 이어서 표시
    # 기존 웹/엔드포인트 결과 개수 및 요약 KPI에는 포함하지 않습니다.
    # ---------------------------------------------------------------
    actual_file_reports = (
        actual_file_reports
        if isinstance(actual_file_reports, list)
        else []
    )

    if actual_file_reports:
        # 기존 상세 결과와 실제 파일 분석 결과 사이 구분
        current_row += 1

        detail_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=9
        )

        actual_section = detail_sheet.cell(
            current_row,
            1,
            "실제 업로드 파일 진단 결과"
        )
        fill(actual_section, HEADER_FILL)
        actual_section.font = Font(
            bold=True,
            color=BLUE,
            size=12
        )
        actual_section.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for col in range(1, 10):
            detail_sheet.cell(
                current_row,
                col
            ).border = DEFAULT_BORDER

        detail_sheet.row_dimensions[
            current_row
        ].height = 28

        current_row += 2

        for file_report in actual_file_reports:
            if not isinstance(file_report, dict):
                continue

            filename = safe_value(
                file_report.get("filename"),
                "알 수 없는 파일"
            )
            file_url = safe_value(
                file_report.get("file_url"),
                "-"
            )
            actual_items = file_report.get(
                "results",
                []
            )

            if not isinstance(actual_items, list):
                actual_items = []

            vulnerable_count = sum(
                1
                for actual_item in actual_items
                if isinstance(actual_item, dict)
                and actual_item.get("status") == "취약"
            )

            # 파일 단위 제목
            detail_sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=9
            )
            file_title_cell = detail_sheet.cell(
                current_row,
                1,
                f"파일: {filename} · 취약 {vulnerable_count}건"
            )
            fill(file_title_cell, "DCE6F1")
            file_title_cell.font = Font(
                bold=True,
                color=BLUE,
                size=11
            )
            file_title_cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

            for col in range(1, 10):
                detail_sheet.cell(
                    current_row,
                    col
                ).border = DEFAULT_BORDER

            detail_sheet.row_dimensions[
                current_row
            ].height = 26

            current_row += 1

            # 실제 업로드 파일 URL
            label_cell(
                detail_sheet,
                current_row,
                "파일 URL"
            )
            detail_sheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=9
            )
            file_url_cell = detail_sheet.cell(
                current_row,
                2,
                file_url
            )
            file_url_cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )
            for col in range(2, 10):
                detail_sheet.cell(
                    current_row,
                    col
                ).border = DEFAULT_BORDER
            detail_sheet.row_dimensions[
                current_row
            ].height = 28

            current_row += 2

            if not actual_items:
                detail_sheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=9
                )
                no_item_cell = detail_sheet.cell(
                    current_row,
                    1,
                    "해당 파일의 정적 분석 결과가 없습니다."
                )
                fill(no_item_cell, "F7F9FB")
                no_item_cell.font = Font(
                    italic=True,
                    color="666666"
                )
                no_item_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                for col in range(1, 10):
                    detail_sheet.cell(
                        current_row,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row
                ].height = 28
                current_row += 2
                continue

            for item_index, actual_item in enumerate(
                actual_items,
                start=1
            ):
                if not isinstance(actual_item, dict):
                    continue

                # 항목 제목
                detail_sheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=9
                )
                title_cell = detail_sheet.cell(
                    current_row,
                    1,
                    f"{item_index:02d}. "
                    f"{safe_value(actual_item.get('vulnerability'))}"
                )
                fill(title_cell, TITLE_FILL)
                title_cell.font = Font(
                    bold=True,
                    color=WHITE,
                    size=11
                )
                title_cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

                for col in range(1, 10):
                    detail_sheet.cell(
                        current_row,
                        col
                    ).border = DEFAULT_BORDER

                detail_sheet.row_dimensions[
                    current_row
                ].height = 27

                # 판정 / 위험도 / 확실성
                status_risk_meta(
                    detail_sheet,
                    current_row + 1,
                    actual_item
                )
                detail_sheet.row_dimensions[
                    current_row + 1
                ].height = 24

                # 진단 대상
                label_cell(
                    detail_sheet,
                    current_row + 2,
                    "진단 대상"
                )
                detail_sheet.merge_cells(
                    start_row=current_row + 2,
                    start_column=2,
                    end_row=current_row + 2,
                    end_column=9
                )
                target_cell = detail_sheet.cell(
                    current_row + 2,
                    2,
                    safe_value(
                        actual_item.get("parameter")
                    )
                )
                target_cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )
                for col in range(2, 10):
                    detail_sheet.cell(
                        current_row + 2,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row + 2
                ].height = 30

                # 테스트 입력값
                label_cell(
                    detail_sheet,
                    current_row + 3,
                    "테스트 입력값"
                )
                detail_sheet.merge_cells(
                    start_row=current_row + 3,
                    start_column=2,
                    end_row=current_row + 3,
                    end_column=9
                )
                payload_cell = detail_sheet.cell(
                    current_row + 3,
                    2,
                    safe_value(
                        actual_item.get("payload")
                    )
                )
                payload_cell.font = Font(
                    name="Consolas",
                    size=9,
                    color="3B3B3B"
                )
                payload_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                for col in range(2, 10):
                    detail_sheet.cell(
                        current_row + 3,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row + 3
                ].height = 44

                # 탐지 내용
                label_cell(
                    detail_sheet,
                    current_row + 4,
                    "탐지 내용"
                )
                detail_sheet.merge_cells(
                    start_row=current_row + 4,
                    start_column=2,
                    end_row=current_row + 4,
                    end_column=9
                )
                evidence_cell = detail_sheet.cell(
                    current_row + 4,
                    2,
                    safe_value(
                        actual_item.get("evidence")
                    )
                )
                evidence_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                for col in range(2, 10):
                    detail_sheet.cell(
                        current_row + 4,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row + 4
                ].height = 46

                # 판단 근거
                label_cell(
                    detail_sheet,
                    current_row + 5,
                    "판단 근거"
                )
                detail_sheet.merge_cells(
                    start_row=current_row + 5,
                    start_column=2,
                    end_row=current_row + 5,
                    end_column=9
                )
                reason_cell = detail_sheet.cell(
                    current_row + 5,
                    2,
                    safe_value(
                        actual_item.get("reason")
                    )
                )
                reason_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                for col in range(2, 10):
                    detail_sheet.cell(
                        current_row + 5,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row + 5
                ].height = 46

                # 대응방안
                label_cell(
                    detail_sheet,
                    current_row + 6,
                    "대응방안"
                )
                detail_sheet.merge_cells(
                    start_row=current_row + 6,
                    start_column=2,
                    end_row=current_row + 6,
                    end_column=9
                )
                rec_cell = detail_sheet.cell(
                    current_row + 6,
                    2,
                    safe_value(
                        actual_item.get("recommendation")
                    )
                )
                rec_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                for col in range(2, 10):
                    detail_sheet.cell(
                        current_row + 6,
                        col
                    ).border = DEFAULT_BORDER
                detail_sheet.row_dimensions[
                    current_row + 6
                ].height = 46

                detail_sheet.row_dimensions[
                    current_row + 7
                ].height = 9

                current_row += 8

            # 파일별 여백
            detail_sheet.row_dimensions[
                current_row
            ].height = 10
            current_row += 1


    detail_sheet.freeze_panes = "A5"

    set_column_widths(
        detail_sheet,
        {
            "A": 18,
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 18
        }
    )

    # ===============================================================
    # 3) 수동-자동 비교
    # ===============================================================
    compare_sheet = workbook.create_sheet(
        "수동-자동 비교"
    )

    style_title(
        compare_sheet,
        "A1:J1",
        "수동 진단 vs 자동 진단 비교"
    )

    rows = normalize_comparison_data(
        comparison_data
    )

    matched_count = sum(
        1 for item in rows
        if safe_value(
            item.get("비교 결과"),
            ""
        ) == "일치"
    )
    mismatch_count = sum(
        1 for item in rows
        if safe_value(
            item.get("비교 결과"),
            ""
        ) == "불일치"
    )
    no_data_count = sum(
        1 for item in rows
        if safe_value(
            item.get("비교 결과"),
            ""
        ) == "비교 데이터 없음"
    )
    comparable_count = (
        matched_count
        + mismatch_count
    )
    match_rate = (
        matched_count / comparable_count
        if comparable_count
        else 0
    )

    # KPI 영역
    kpi = [
        ("비교 가능", comparable_count),
        ("일치", matched_count),
        ("불일치", mismatch_count),
        ("일치율", match_rate)
    ]

    for idx, (label, value) in enumerate(
        kpi
    ):
        start_col = 1 + idx * 2

        compare_sheet.merge_cells(
            start_row=3,
            start_column=start_col,
            end_row=3,
            end_column=start_col + 1
        )
        compare_sheet.merge_cells(
            start_row=4,
            start_column=start_col,
            end_row=4,
            end_column=start_col + 1
        )

        header_cell = compare_sheet.cell(
            3,
            start_col,
            label
        )
        fill(header_cell, HEADER_FILL)
        header_cell.font = Font(
            bold=True
        )
        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        value_cell = compare_sheet.cell(
            4,
            start_col,
            value
        )
        value_cell.font = Font(
            bold=True,
            size=14,
            color=BLUE
        )
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for row_idx in [3, 4]:
            for col_idx in range(
                start_col,
                start_col + 2
            ):
                compare_sheet.cell(
                    row_idx,
                    col_idx
                ).border = DEFAULT_BORDER

        if label == "일치율":
            value_cell.number_format = "0.0%"

    if no_data_count:
        compare_sheet.merge_cells(
            "I3:J4"
        )
        info = compare_sheet["I3"]
        info.value = (
            f"비교 데이터 없음\n"
            f"{no_data_count}건"
        )
        fill(info, "F3F6F9")
        info.font = Font(
            bold=True,
            color="666666"
        )
        info.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        set_all_border(
            compare_sheet,
            "I3:J4"
        )

    compare_headers = [
        "No.",
        "취약점",
        "수동 진단",
        "자동 진단",
        "비교 결과",
        "비교 사유"
    ]

    for index, value in enumerate(
        compare_headers,
        start=1
    ):
        compare_sheet.cell(
            row=7,
            column=index,
            value=value
        )

    style_header_row(
        compare_sheet,
        7,
        1,
        6
    )

    current_row = 8

    if rows:
        for idx, item in enumerate(
            rows,
            start=1
        ):
            values = [
                idx,
                safe_value(
                    item.get("취약점")
                ),
                safe_value(
                    item.get("수동 진단")
                ),
                safe_value(
                    item.get("자동 진단")
                ),
                safe_value(
                    item.get("비교 결과")
                ),
                safe_value(
                    item.get("비교 사유")
                )
            ]

            for column, value in enumerate(
                values,
                start=1
            ):
                cell = compare_sheet.cell(
                    row=current_row,
                    column=column,
                    value=value
                )
                cell.border = DEFAULT_BORDER
                cell.alignment = Alignment(
                    horizontal=(
                        "center"
                        if column in [1, 3, 4, 5]
                        else "left"
                    ),
                    vertical="center",
                    wrap_text=True
                )

            apply_status_style(
                compare_sheet.cell(
                    current_row,
                    3
                )
            )
            apply_status_style(
                compare_sheet.cell(
                    current_row,
                    4
                )
            )

            result_cell = compare_sheet.cell(
                current_row,
                5
            )

            if result_cell.value == "일치":
                fill(
                    result_cell,
                    GREEN_FILL
                )
                result_cell.font = Font(
                    bold=True,
                    color=GREEN
                )

            elif result_cell.value == "불일치":
                fill(
                    result_cell,
                    RED_FILL
                )
                result_cell.font = Font(
                    bold=True,
                    color=RED
                )

            compare_sheet.row_dimensions[
                current_row
            ].height = 36

            current_row += 1

    else:
        compare_sheet.merge_cells(
            "A8:F10"
        )
        no_data_cell = compare_sheet["A8"]
        no_data_cell.value = (
            "업로드된 수동 진단 결과가 없어 "
            "현재 비교 가능한 데이터가 없습니다."
        )
        no_data_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        no_data_cell.font = Font(
            italic=True,
            color="666666"
        )
        fill(
            no_data_cell,
            "F7F9FB"
        )
        set_all_border(
            compare_sheet,
            "A8:F10"
        )

    compare_sheet.freeze_panes = "A8"

    set_column_widths(
        compare_sheet,
        {
            "A": 7,
            "B": 40,
            "C": 14,
            "D": 14,
            "E": 18,
            "F": 68,
            "G": 5,
            "H": 5,
            "I": 16,
            "J": 16
        }
    )

    # ===============================================================
    # 4) AI 보조 분석
    # ===============================================================
    ai_sheet = workbook.create_sheet(
        "AI 보조 분석"
    )

    style_title(
        ai_sheet,
        "A1:D1",
        "AI 기반 보조 분석 결과"
    )

    current_row = 3

    if ai_analysis:
        ai_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4
        )
        section = ai_sheet.cell(
            current_row,
            1,
            "OpenAI 종합 분석"
        )
        fill(section, HEADER_FILL)
        section.font = Font(
            bold=True,
            color=BLUE
        )
        section.alignment = Alignment(
            vertical="center"
        )
        set_all_border(
            ai_sheet,
            f"A{current_row}:D{current_row}"
        )
        ai_sheet.row_dimensions[
            current_row
        ].height = 24

        current_row += 1

        cleaned_ai_text = clean_ai_report_text(
            ai_analysis
        )

        # 빈 줄을 제외한 실제 문장/제목/목록 단위로 한 행씩 배치
        ai_lines = [
            line.strip()
            for line in cleaned_ai_text.splitlines()
            if line.strip()
        ]

        if not ai_lines:
            ai_lines = [
                "AI 종합 분석 결과가 없습니다."
            ]

        for line in ai_lines:
            # 제목처럼 보이는 줄 판별
            is_heading = bool(
                re.match(
                    r"^(?:\d+\.\s+|[가-힣A-Za-z ]{2,25}:?$)",
                    line
                )
            ) and not line.startswith(("-", "•"))

            ai_sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=4
            )

            content = ai_sheet.cell(
                current_row,
                1,
                line
            )

            if is_heading:
                fill(content, "F3F6F9")
                content.font = Font(
                    bold=True,
                    color=BLUE,
                    size=10
                )
            else:
                content.font = Font(
                    color=DARK,
                    size=10
                )

            content.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

            set_all_border(
                ai_sheet,
                f"A{current_row}:D{current_row}"
            )

            # 한 행이 지나치게 높아지지 않으면서
            # 한국어 긴 문장도 겹치지 않도록 높이 계산
            estimated_lines = max(
                1,
                (len(line) // 58) + 1
            )

            ai_sheet.row_dimensions[
                current_row
            ].height = min(
                72,
                max(
                    22,
                    estimated_lines * 18
                )
            )

            current_row += 1

        current_row += 1

    if hf_result:
        ai_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4
        )
        section = ai_sheet.cell(
            current_row,
            1,
            "Hugging Face SQL Injection 보조 분석"
        )
        fill(section, HEADER_FILL)
        section.font = Font(
            bold=True,
            color=BLUE
        )
        set_all_border(
            ai_sheet,
            f"A{current_row}:D{current_row}"
        )

        current_row += 1

        hf_rows = [
            (
                "진단 대상",
                safe_value(
                    hf_result.get("parameter")
                )
            ),
            (
                "분석 입력값",
                safe_value(
                    hf_result.get("payload")
                )
            ),
            (
                "AI 판정",
                safe_value(
                    hf_result.get("status")
                )
            ),
            (
                "모델 확신도",
                f"{hf_result.get('score', 0) * 100:.2f}%"
            )
        ]

        for label, value in hf_rows:
            ai_sheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=4
            )

            label_cell_obj = ai_sheet.cell(
                current_row,
                1,
                label
            )
            fill(
                label_cell_obj,
                SUB_FILL
            )
            label_cell_obj.font = Font(
                bold=True,
                color=BLUE
            )
            label_cell_obj.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            label_cell_obj.border = DEFAULT_BORDER

            value_cell = ai_sheet.cell(
                current_row,
                2,
                value
            )
            value_cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

            for col in range(2, 5):
                ai_sheet.cell(
                    current_row,
                    col
                ).border = DEFAULT_BORDER

            ai_sheet.row_dimensions[
                current_row
            ].height = 30

            current_row += 1

        current_row += 1

    if not ai_analysis and not hf_result:
        ai_sheet.merge_cells(
            "A3:D5"
        )
        info = ai_sheet["A3"]
        info.value = (
            "현재 저장된 AI 분석 결과가 없습니다."
        )
        info.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        fill(info, "F7F9FB")
        set_all_border(
            ai_sheet,
            "A3:D5"
        )

    set_column_widths(
        ai_sheet,
        {
            "A": 24,
            "B": 28,
            "C": 28,
            "D": 28
        }
    )

    ai_sheet.freeze_panes = "A3"

    # ===============================================================
    # 전체 시트 공통
    # ===============================================================
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if isinstance(
                        cell.value,
                        str
                    ):
                        cell.value = (
                            sanitize_excel_text(
                                cell.value
                            )
                        )

                    cell.alignment = Alignment(
                        horizontal=(
                            cell.alignment.horizontal
                            or "left"
                        ),
                        vertical=(
                            cell.alignment.vertical
                            or "top"
                        ),
                        wrap_text=True
                    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
